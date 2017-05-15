# TOOL korp-kwic-xlsx-frame.py: "DEPRECATED Korp KWIC to XLSX Frame"
# (DEPRECATED Flatten a JSON concordance in Excel form, with header.)
# INPUT kwic.json TYPE GENERIC
# OUTPUT kwic.xlsx
# RUNTIME python3

'''Turn JSON format KWIC concordance from Korp API to a flat, headed
   Excel spreadsheet. Use the attribute names from the input KWIC for
   the output columns. Repeat structural attributes of a hit for each
   token. -- Depends on openpyxl.

'''

import json, os, sys
from itertools import chain, count
from openpyxl import Workbook

sys.path.append(os.path.join(chipster_module_path, "python"))
import lib_names as names

names.output('kwic.xlsx', names.replace('kwic.json', '.xlsx'))

with open('kwic.json', encoding = 'utf-8') as f:
    data = json.load(f)

kwic = data['kwic']

# lead sentence/token determines which attributes,
# any CoNLL equivalents first,
# then match and corpus,
# then a couple of our own,
# then other positionals lexicographically
# then structurals lexicographically

lead = kwic[0]['tokens'][0] # lead token
head = [ key
         for key in 'ref word lemma pos msd dephead deprel'.split()
         if key in lead ]

what = 'match_start', 'match_end', 'corpus' # hope them is free!

rest = [ key for key in sorted(lead) if key not in head ]
meta = sorted(kwic[0]['structs'])

# is there something wrong about the match start and end?
# also, should they be just 0/1 anyway?
# also, should there also be a hit counter?

# Begins!

book = Workbook()
sheet = book.active

numbers = count(start = 1)
headrow = next(numbers)
for k, name in enumerate(chain(head, what, rest, meta), start = 1):
    sheet.cell(row = headrow, column = k, value = name) # really

for hit in kwic:
    for token in hit['tokens']:
        row = next(numbers)
        for k, value in enumerate(chain((token[key] for key in head),
                                        (hit['match']['start'],
                                         hit['match']['end'],
                                         hit['corpus']),
                                        (token[key] for key in rest),
                                        (hit['structs'][key] for key in meta)),
                                  start = 1):
            sheet.cell(row = row, column = k, value = value) # really

book.save('kwic.tmp')
os.rename('kwic.tmp', 'kwic.xlsx')
