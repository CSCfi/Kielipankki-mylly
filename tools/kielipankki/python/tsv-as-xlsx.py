# TOOL tsv-as-xlsx.py: "TSV as XLSX"
# (TSV as Microsoft Excel spreadsheet)
# INPUT table.tsv: "TSV file" TYPE GENERIC
# OUTPUT table.xlsx
# IMAGE comp-16.04-mylly
# RUNTIME python3

import os, sys
from itertools import chain, count
from openpyxl import Workbook

sys.path.append(os.path.join(chipster_module_path, "python"))
from lib_names2 import base, name

name('table.xlsx', base('table.tsv', '*.tsv'),
     ext = 'xlsx')

# Begins!

book = Workbook()
sheet = book.active

with open('table.tsv', encoding = 'utf-8') as tsv:
    # header
    head = next(tsv).rstrip('\n').split('\t')
    for k, name in enumerate(head, start = 1):
        sheet.cell(row = 1, column = k, value = name) # really
    # records
    for r, line in enumerate(tsv, start = 2):
        record = line.rstrip('\n').split('\t')
        for k, value in enumerate(record, start = 1):
            sheet.cell(row = r, column = k, value = value) # really

book.save('table.tmp')
os.rename('table.tmp', 'table.xlsx')
